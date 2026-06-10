import sys
import pandas as pd
import warnings
import yaml
import os
import json
import openpyxl
from datetime import datetime
# 数据匹配
sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore', message='Workbook contains no default style')

# 读取配置
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

# 收益列名：优先使用yaml配置，否则根据当前月份自动生成
_config_range = config.get('dataRange', '')
if _config_range:
    REVENUE_COL = _config_range
else:
    _current_month = datetime.now().month
    REVENUE_COL = f'1-{_current_month - 1}月收益'

# 数据年份：优先使用yaml配置，否则根据当前年份自动生成
_config_year = config.get('dataYear', '')
if _config_year:
    DATA_YEAR = _config_year
else:
    DATA_YEAR = str(datetime.now().year)[2:]
YEAR_REVENUE_COL = f'{DATA_YEAR}年收益'
CURRENT_CONTRACT_PREFIX = f'C{2000 + int(DATA_YEAR) + 1}'

start_date = config['date_range']['start_date']
end_date = config['date_range']['end_date']
start_str = start_date.replace('-', '')
end_str = end_date.replace('-', '')


def main():
    # 文件路径：支持 workflow 时间戳
    ts = os.environ.get('WORKFLOW_TIMESTAMP', '')
    suffix = f'_{ts}' if ts else ''
    input_file = f'中间数据/合同汇总{start_str}-{end_str}{suffix}.xlsx'
    output_file = input_file  # 原地更新

    # 先用 openpyxl 读取底纹，过滤掉带底纹的行（红色/黄色=未匹配）
    wb_check = openpyxl.load_workbook(input_file)
    ws_check = wb_check['合同汇总'] if '合同汇总' in wb_check.sheetnames else wb_check.active
    shaded_rows = set()  # 带底纹的行号（1-indexed）
    for row_idx in range(2, ws_check.max_row + 1):
        cell = ws_check.cell(row=row_idx, column=1)
        fill = cell.fill
        if fill and fill.fill_type == 'solid' and fill.start_color and fill.start_color.rgb:
            rgb = str(fill.start_color.rgb)
            # FF0000=红色, FFFF00=黄色
            if rgb in ('FF0000', '00FF0000', 'FFFF00', '00FFFF00'):
                shaded_rows.add(row_idx)
    wb_check.close()

    df = pd.read_excel(input_file)
    print(f"读取: {input_file}, 共 {len(df)} 条记录")
    if shaded_rows:
        df = df.drop(index=[r - 2 for r in shaded_rows if r - 2 < len(df)]).reset_index(drop=True)
        print(f"过滤带底纹行: {len(shaded_rows)} 行, 剩余: {len(df)} 行")

    # 从合同编号解析年份，当年合同为新开专区
    contract_year = df['合同编号'].str[1:5].astype(int)
    current_year = int(f'20{DATA_YEAR}')
    is_new = contract_year == current_year
    df['专属属性'] = is_new.map({True: '新开', False: '历史'})

    new_count = is_new.sum()
    print(f"新开: {new_count} 个, 历史: {len(df) - new_count} 个")

    # 预算使用率 = 预算使用 ÷ 核定总额
    df['预算使用率'] = df.apply(
        lambda row: round(row['预算使用'] / row['核定总额'], 4)
        if pd.notna(row.get('预算使用')) and pd.notna(row.get('核定总额')) and row['核定总额'] > 0
        else 0,
        axis=1
    )

    # 确保输出时带有百分号
    if df['预算使用率'].dtype == 'float64':
        df['预算使用率'] = df['预算使用率'].apply(lambda x: f"{round(x * 100, 2)}%" if pd.notna(x) else "0%")
    else:
        df['预算使用率'] = df['预算使用率'].apply(
            lambda x: str(x) if pd.notna(x) and str(x).endswith('%') else (f"{x}%" if pd.notna(x) else "0%")
        )

    # 计算收入成本比 = 实际运营收益 ÷ 已使用预算
    # 实际运营收益 = REVENUE_COL列, 已使用预算 = 预算使用列
    # 列为空时按0计算
    revenue_col = REVENUE_COL if REVENUE_COL in df.columns else YEAR_REVENUE_COL
    print(f"收益列名: {revenue_col}")
    df[revenue_col] = df[revenue_col].fillna(0)
    df['收入成本比'] = df.apply(
        lambda row: round(row[revenue_col] / row['预算使用'], 2) if row['预算使用'] > 0 else 0,
        axis=1
    )

    print(f"预算使用率计算完成")
    print(f"收入成本比计算完成")

    # 分类函数：判断每行属于哪个颜色
    def classify_row(row):
        # 将百分比字符串转换为数值
        budget_rate_str = row['预算使用率']
        budget_rate = float(budget_rate_str.replace('%', ''))
        cost_ratio = row['收入成本比']
        revenue_1_4 = row[revenue_col] if pd.notna(row[revenue_col]) else 0
        revenue_25 = row[YEAR_REVENUE_COL] if pd.notna(row[YEAR_REVENUE_COL]) else 0

        # 绿色（良性运营）：收入成本比≥1.72
        if cost_ratio >= 1.72:
            return '绿色'

        # 红色（严重问题）
        # 预算使用率≥100% 且 实际收益 = 0
        if budget_rate >= 100 and revenue_1_4 == 0:
            return '红色'

        # 橙色（需改进）
        # 1. 50%≤预算使用率＜100% 且 收入成本比＜1
        if 50 <= budget_rate < 100 and cost_ratio < 1:
            return '橙色'
        # 2. 预算使用率≥100%、实际收益＞0且收入成本比＜1
        if budget_rate >= 100 and revenue_1_4 > 0 and cost_ratio < 1:
            return '橙色'

        # 黄色（持续关注）：30%≤预算使用率＜50% 且 实际收益 = 0
        if 30 <= budget_rate < 50 and revenue_1_4 == 0:
            return '黄色'

        # 不符合条件的返回空
        return ''

    # 添加颜色分类列
    df['颜色分类'] = df.apply(classify_row, axis=1)

    # 统计各颜色数量
    color_counts = df['颜色分类'].value_counts()
    print(f"\n颜色分类统计:")
    for color in ['红色', '橙色', '黄色', '绿色']:
        count = color_counts.get(color, 0)
        print(f"  {color}: {count} 个")
    print(f"  不符合条件: {len(df[df['颜色分类'] == ''])} 个")

    # 用openpyxl原地更新，保留底纹等格式
    wb = openpyxl.load_workbook(output_file)
    ws = wb['合同汇总'] if '合同汇总' in wb.sheetnames else wb.active

    # 找到已有列
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # 需要新增的列
    new_cols = ['专属属性', '预算使用率', '收入成本比', '颜色分类']
    for col_name in new_cols:
        if col_name not in headers:
            next_col = ws.max_column + 1
            ws.cell(row=1, column=next_col, value=col_name)
            headers[col_name] = next_col

    # 写入数据（只写非底纹行）
    # 构建 Excel 行号到 df 索引的映射
    clean_excel_rows = [r for r in range(2, ws.max_row + 1) if r not in shaded_rows]
    for df_idx, excel_row in enumerate(clean_excel_rows):
        if df_idx >= len(df):
            break
        for col_name in new_cols:
            col_idx = headers[col_name]
            val = df.iloc[df_idx][col_name]
            ws.cell(row=excel_row, column=col_idx, value=str(val) if pd.notna(val) else '')

    # 各颜色分类sheet
    for color in ['红色', '橙色', '黄色', '绿色']:
        color_df = df[df['颜色分类'] == color]
        if color_df.empty:
            continue
        sheet_name = f'{color}专区'
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_color = wb.create_sheet(sheet_name)
        # 写表头
        for c, col_name in enumerate(df.columns, 1):
            ws_color.cell(row=1, column=c, value=col_name)
        # 写数据
        for r, (_, row) in enumerate(color_df.iterrows(), 2):
            for c, col_name in enumerate(df.columns, 1):
                val = row[col_name]
                ws_color.cell(row=r, column=c, value=str(val) if pd.notna(val) else '')

    wb.save(output_file)

    print(f"\n导出完成: {output_file}")
    print(f"包含sheet: 合同汇总, 红色专区, 橙色专区, 黄色专区, 绿色专区")

    # 计算统计数据并保存到JSON
    stats = {
        "预算使用总额": float(round(df['预算使用'].sum(), 2)),
        "实际运营收益": float(round(df[revenue_col].sum(), 2)),
        "有成本投入专区总数": int(df['预算使用'].count()),
        "红色专区个数": int(color_counts.get('红色', 0)),
        "橙色专区个数": int(color_counts.get('橙色', 0)),
        "黄色专区个数": int(color_counts.get('黄色', 0)),
        "绿色专区个数": int(color_counts.get('绿色', 0))
    }

    json_file = f'中间数据/统计数据{start_str}-{end_str}{suffix}.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)

    print(f"\n统计数据已保存到: {json_file}")
    for key, value in stats.items():
        print(f"  {key}: {value}")


if __name__ == '__main__':
    main()
