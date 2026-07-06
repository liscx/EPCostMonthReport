import sys
import pandas as pd
import warnings
import yaml
import os
import re
from datetime import datetime
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore', message='Workbook contains no default style')

# 读取配置生成文件名
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

# 数据年份：优先使用yaml配置，否则根据当前年份自动生成
_config_year = config.get('dataYear', '')
if _config_year:
    DATA_YEAR_SHORT = _config_year
else:
    DATA_YEAR_SHORT = str(datetime.now().year)[2:]
YEAR_REVENUE_COL = f'{DATA_YEAR_SHORT}年收益'
PRIOR_YEAR_REVENUE_COL = f'{int(DATA_YEAR_SHORT) - 1}年收益'

start_date = config['date_range']['start_date']
end_date = config['date_range']['end_date']
start_str = start_date.replace('-', '')
end_str = end_date.replace('-', '')

SOURCE_FILE = f'源数据/export/projExport{start_str}-{end_str}.xlsx'
REF_FILE = '源数据/合同汇总表.xlsx'
REF_FILE_OLD = '源数据/新点电子交易专区&项目跟进表.xlsx'

# 按列索引定位 projExport.xlsx
COL_CONTRACT = 20   # 合同编号
COL_COST = 12       # 实际人工成本
COL_BUDGET = 23     # 任务预算使用


def load_contracts_from_ejy():
    """从 projExport.xlsx 提取C开头的合同编号及汇总数据，返回 {合同编号: {实际人工成本, 任务预算使用}} 字典"""
    xls = pd.ExcelFile(SOURCE_FILE)
    print(f"源文件: {SOURCE_FILE}, 共 {len(xls.sheet_names)} 个Sheet")

    all_data = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
        if df.empty or len(df.columns) <= max(COL_CONTRACT, COL_COST, COL_BUDGET):
            print(f"  跳过空Sheet: {sheet_name}")
            continue

        contract = df.iloc[:, COL_CONTRACT]
        cost = pd.to_numeric(df.iloc[:, COL_COST], errors='coerce').fillna(0)
        budget = pd.to_numeric(df.iloc[:, COL_BUDGET], errors='coerce').fillna(0)

        temp = pd.DataFrame({
            '合同编号': contract,
            '实际人工成本': cost,
            '任务预算使用': budget,
        })
        temp = temp[temp['合同编号'].notna() & (temp['合同编号'] != '')]
        temp = temp[temp['合同编号'].astype(str).str.startswith('C')]
        if not temp.empty:
            all_data.append(temp)
            print(f"  {sheet_name}: {len(temp)} 条记录")

    if not all_data:
        print("未找到任何数据")
        return None

    merged = pd.concat(all_data, ignore_index=True)
    agg = merged.groupby('合同编号', as_index=False).agg(
        实际人工成本=('实际人工成本', 'sum'),
        任务预算使用=('任务预算使用', 'sum'),
    )
    agg['实际人工成本'] = agg['实际人工成本'].round(2)
    agg['任务预算使用'] = agg['任务预算使用'].round(2)

    cost_dict = {}
    for _, r in agg.iterrows():
        cost_dict[str(r['合同编号'])] = {
            '实际人工成本': r['实际人工成本'],
            '任务预算使用': r['任务预算使用'],
        }
    print(f"\n去重后合同数: {len(cost_dict)}")
    return cost_dict


def load_contract_summary():
    """从合同汇总表加载数据，返回 DataFrame"""
    print(f"\n读取合同汇总表: {REF_FILE}")
    df = pd.read_excel(REF_FILE)
    print(f"  行数: {len(df)}, 列数: {len(df.columns)}")
    return df


def parse_contract_ids(contract_id):
    """解析合同编号，处理各种格式：换行符分隔、直接连接等"""
    # 先按换行符分割
    lines = [c.strip() for c in contract_id.split('\n') if c.strip()]

    all_cids = []
    for line in lines:
        # 检查是否是多个合同编号直接连接的情况（如 C2024120035C2026020017）
        matches = re.findall(r'C\d{10}', line)
        if matches:
            all_cids.extend(matches)
        elif line.startswith('C') and len(line) == 12:
            all_cids.append(line)

    return all_cids


def match_contracts(cost_dict, summary_df):
    """按照合同汇总表顺序匹配，返回 (matched, unmatched_red)
    - matched: projExport和合同汇总表都能匹配到的行
    - unmatched_red: 合同汇总表有但projExport没有的行（红色底纹）
    """
    matched_rows = []
    unmatched_red_rows = []

    # 从旧参考文件构建合同编号到专区码的映射
    df_zone = pd.read_excel(REF_FILE_OLD, sheet_name='专区管控表', header=0, usecols=[0, 1, 2, 5, 11, 15])
    df_zone.columns = ['ref_contract', '专区码', '专区名称', '分公司', '商务', '专区状态']
    df_zone = df_zone.dropna(subset=['ref_contract'])
    df_zone['ref_contract'] = df_zone['ref_contract'].astype(str)

    # 先收集每个合同编号的所有记录（处理重复情况）
    cid_records = {}
    for _, ref_row in df_zone.iterrows():
        ref_contract = str(ref_row['ref_contract'])
        zone_code = str(ref_row['专区码']) if pd.notna(ref_row['专区码']) else ''
        zone_status = str(ref_row['专区状态']) if pd.notna(ref_row['专区状态']) else ''
        for cid in ref_contract.split('\n'):
            cid = cid.strip()
            if cid and cid.startswith('C'):
                if cid not in cid_records:
                    cid_records[cid] = []
                cid_records[cid].append({'专区码': zone_code, '专区状态': zone_status})

    # 对重复合同：优先取非已下线的记录
    cid_to_zone = {}
    duplicates_marked = []
    for cid, records in cid_records.items():
        if len(records) == 1:
            cid_to_zone[cid] = records[0]['专区码']
        else:
            # 有重复，优先取非已下线
            active = [r for r in records if r['专区状态'] != '已下线']
            if active:
                cid_to_zone[cid] = active[0]['专区码']
            else:
                cid_to_zone[cid] = records[0]['专区码']
            duplicates_marked.append(cid)

    if duplicates_marked:
        print(f"  [标记] 发现重复合同（已自动处理）: {len(duplicates_marked)} 个")
        for cid in duplicates_marked:
            print(f"    {cid}: {cid_records[cid]}")

    print(f"  旧参考文件专区码映射数量: {len(cid_to_zone)}")

    # 按合同汇总表顺序遍历
    for _, row in summary_df.iterrows():
        contract_id = str(row['合同编号']).strip() if pd.notna(row['合同编号']) else ''
        if not contract_id:
            continue

        cids = parse_contract_ids(contract_id)
        if not cids:
            # 无法解析，作为红色（合同汇总表有，projExport无）
            unmatched_red_rows.append({
                '合同编号': contract_id,
                '实际人工成本': '',
                '预算使用': '',
                '分公司': '',
                '专区名称': str(row['专区名称']) if pd.notna(row.get('专区名称', '')) else '',
                '商务': str(row['商务']) if pd.notna(row.get('商务', '')) else '',
                '专区码': '',
                '核定总额': row.get('核定总额', ''),
                PRIOR_YEAR_REVENUE_COL: row.get(PRIOR_YEAR_REVENUE_COL, ''),
                YEAR_REVENUE_COL: row.get(YEAR_REVENUE_COL, ''),
                '核定总额计算规则': str(row.get('核定总额计算规则', '')) if pd.notna(row.get('核定总额计算规则', '')) else '',
            })
            continue

        # 在 projExport 中查找
        matched_cids_in_row = []
        total_budget = 0
        for cid in cids:
            if cid in cost_dict:
                matched_cids_in_row.append(cid)
                total_budget += cost_dict[cid]['任务预算使用']

        # 获取分公司
        branch = ''
        if '分公司' in row.index:
            branch = str(row['分公司']) if pd.notna(row['分公司']) else ''
        elif '省份（已按新组织调整）' in row.index:
            branch = str(row['省份（已按新组织调整）']) if pd.notna(row['省份（已按新组织调整）']) else ''

        # 从旧参考文件获取专区码
        zone_code = ''
        for cid in cids:
            if cid in cid_to_zone:
                zone_code = cid_to_zone[cid]
                break

        if matched_cids_in_row:
            # 匹配成功
            matched_rows.append({
                '合同编号': '\n'.join(matched_cids_in_row),
                '实际人工成本': sum(cost_dict[cid]['实际人工成本'] for cid in matched_cids_in_row),
                '预算使用': round(total_budget, 2),
                '分公司': branch,
                '专区名称': str(row['专区名称']) if pd.notna(row.get('专区名称', '')) else '',
                '商务': str(row['商务']) if pd.notna(row.get('商务', '')) else '',
                '专区码': zone_code,
                '核定总额': row.get('核定总额', ''),
                PRIOR_YEAR_REVENUE_COL: row.get(PRIOR_YEAR_REVENUE_COL, ''),
                YEAR_REVENUE_COL: row.get(YEAR_REVENUE_COL, ''),
                '核定总额计算规则': str(row.get('核定总额计算规则', '')) if pd.notna(row.get('核定总额计算规则', '')) else '',
            })
        else:
            # 合同汇总表有，projExport没有 -> 红色底纹
            unmatched_red_rows.append({
                '合同编号': contract_id,
                '实际人工成本': '',
                '预算使用': '',
                '分公司': branch,
                '专区名称': str(row['专区名称']) if pd.notna(row.get('专区名称', '')) else '',
                '商务': str(row['商务']) if pd.notna(row.get('商务', '')) else '',
                '专区码': zone_code,
                '核定总额': row.get('核定总额', ''),
                PRIOR_YEAR_REVENUE_COL: row.get(PRIOR_YEAR_REVENUE_COL, ''),
                YEAR_REVENUE_COL: row.get(YEAR_REVENUE_COL, ''),
                '核定总额计算规则': str(row.get('核定总额计算规则', '')) if pd.notna(row.get('核定总额计算规则', '')) else '',
            })

    return matched_rows, unmatched_red_rows


def match_unmatched_ejy(cost_dict, matched_in_summary, summary_df):
    """处理 projExport 中未在合同汇总表匹配到的合同（黄色底纹），尝试从旧参考文件补充信息"""
    print(f"\n处理 projExport 未匹配合同，参考旧文件: {REF_FILE_OLD}")

    # 从旧参考文件构建合同编号到信息的映射
    df_zone = pd.read_excel(REF_FILE_OLD, sheet_name='专区管控表', header=0, usecols=[0, 1, 2, 5, 11, 15])
    df_zone.columns = ['ref_contract', '专区码', '专区名称', '分公司', '商务', '专区状态']
    df_zone = df_zone.dropna(subset=['ref_contract'])
    df_zone['ref_contract'] = df_zone['ref_contract'].astype(str)

    # 先收集每个合同编号的所有记录
    cid_all_records = {}
    for _, ref_row in df_zone.iterrows():
        ref_contract = str(ref_row['ref_contract'])
        zone_code = str(ref_row['专区码']) if pd.notna(ref_row['专区码']) else ''
        zone_status = str(ref_row['专区状态']) if pd.notna(ref_row['专区状态']) else ''
        info = {
            '专区码': zone_code,
            '专区名称': str(ref_row['专区名称']) if pd.notna(ref_row['专区名称']) else '',
            '分公司': str(ref_row['分公司']) if pd.notna(ref_row['分公司']) else '',
            '商务': str(ref_row['商务']) if pd.notna(ref_row['商务']) else '',
            '专区状态': zone_status,
        }
        for cid in ref_contract.split('\n'):
            cid = cid.strip()
            if cid and cid.startswith('C'):
                if cid not in cid_all_records:
                    cid_all_records[cid] = []
                cid_all_records[cid].append(info)

    # 对重复合同：优先取非已下线的记录
    cid_to_info = {}
    for cid, records in cid_all_records.items():
        if len(records) == 1:
            cid_to_info[cid] = records[0]
        else:
            active = [r for r in records if r['专区状态'] != '已下线']
            if active:
                cid_to_info[cid] = active[0]
            else:
                cid_to_info[cid] = records[0]

    # 收集合同汇总表中已匹配的合同编号集合（用于排除）
    summary_cids = set()
    for _, row in summary_df.iterrows():
        contract_id = str(row['合同编号']).strip() if pd.notna(row['合同编号']) else ''
        if contract_id:
            cids = parse_contract_ids(contract_id)
            summary_cids.update(cids)

    yellow_rows = []
    for cid, costs in cost_dict.items():
        if cid in summary_cids:
            continue  # 已在合同汇总表中匹配过，跳过

        info = cid_to_info.get(cid, {})
        yellow_rows.append({
            '合同编号': cid,
            '实际人工成本': costs['实际人工成本'],
            '预算使用': costs['任务预算使用'],
            '分公司': info.get('分公司', ''),
            '专区名称': info.get('专区名称', ''),
            '商务': info.get('商务', ''),
            '专区码': info.get('专区码', ''),
            '核定总额': '',
            PRIOR_YEAR_REVENUE_COL: '',
            YEAR_REVENUE_COL: '',
            '核定总额计算规则': '',
        })

    print(f"  projExport 未匹配: {len(yellow_rows)} 行")
    return yellow_rows


def main():
    # 文件路径：支持 workflow 时间戳
    ts = os.environ.get('WORKFLOW_TIMESTAMP', '')
    suffix = f'_{ts}' if ts else ''
    output_file = f'中间数据/合同汇总{start_str}-{end_str}{suffix}.xlsx'

    # 加载 projExport 数据（主数据源）
    cost_dict = load_contracts_from_ejy()
    if cost_dict is None:
        return

    # 加载合同汇总表（参考数据）
    summary_df = load_contract_summary()

    # 第一步：按合同汇总表顺序匹配
    # matched_rows: 两边都能匹配到
    # unmatched_red_rows: 合同汇总表有，projExport没有（红色底纹）
    matched_rows, unmatched_red_rows = match_contracts(cost_dict, summary_df)

    print(f"\n合同汇总表匹配结果:")
    print(f"  匹配成功（绿色）: {len(matched_rows)} 行")
    print(f"  合同汇总表未匹配（红色）: {len(unmatched_red_rows)} 行")

    # 第二步：处理 projExport 中未在合同汇总表匹配到的合同（黄色底纹）
    yellow_rows = match_unmatched_ejy(cost_dict, set(), summary_df)

    # 合并：先按合同汇总表顺序（匹配+红色），再追加黄色
    all_rows = matched_rows + unmatched_red_rows + yellow_rows
    result = pd.DataFrame(all_rows)

    print(f"\n最终结果:")
    print(f"  匹配成功（绿色）: {len(matched_rows)} 行")
    print(f"  合同汇总表未匹配（红色）: {len(unmatched_red_rows)} 行")
    print(f"  projExport 未匹配（黄色）: {len(yellow_rows)} 行")

    # 写入Excel并添加底纹
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name='合同汇总')

        workbook = writer.book
        worksheet = workbook['合同汇总']

        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        matched_end = len(matched_rows)
        red_end = matched_end + len(unmatched_red_rows)

        # 红色：合同汇总表有，projExport没有
        for row_idx in range(matched_end + 2, red_end + 2):
            for col_idx in range(1, len(result.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

        # 黄色：projExport有，合同汇总表没有
        for row_idx in range(red_end + 2, len(result) + 2):
            for col_idx in range(1, len(result.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

    print(f"\n导出完成: {output_file}")
    print(f"总行数: {len(result)}")


if __name__ == '__main__':
    main()
