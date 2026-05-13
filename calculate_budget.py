import pandas as pd
import sys
import yaml
import os
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# 读取配置
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

# 数据年份：优先使用yaml配置，否则根据当前年份自动生成
_config_year = config.get('dataYear', '')
if _config_year:
    DATA_YEAR_SHORT = _config_year
else:
    DATA_YEAR_SHORT = str(datetime.now().year)[2:]
DATA_YEAR = int(f'20{DATA_YEAR_SHORT}')
NEXT_YEAR = DATA_YEAR + 1
YEAR_REVENUE_COL = f'{DATA_YEAR_SHORT}年收益'

start_date = config['date_range']['start_date']
end_date = config['date_range']['end_date']
start_str = start_date.replace('-', '')
end_str = end_date.replace('-', '')




def calc_quota(row):
    contract_id = str(row['合同编号'])
    revenue = row[YEAR_REVENUE_COL]

    year = int(contract_id[1:5])

    # 1. 当前年+1：新开专区，固定50000
    if year == NEXT_YEAR:
        return 50000, f'{NEXT_YEAR}年新开专区，核定总额=50000'

    # 2. 当前年10月之后，收益为空或为0，按50000算
    if year == DATA_YEAR:
        month = int(contract_id[5:7])
        if month >= 10 and (pd.isna(revenue) or revenue == 0):
            return 50000, f'{DATA_YEAR}年{month}月开设，收益为空或0，新开专区，核定总额=50000'

    # 3. 当前年10月之前收益为空或为0，保底32000
    if year == DATA_YEAR and month < 10 and (pd.isna(revenue) or revenue == 0):
        return 32000, f'{DATA_YEAR}年{month}月开设，收益为空或0，保底32000'

    # 4. 收益为空或为0（当前年之前），保底32000
    if pd.isna(revenue) or revenue == 0:
        return 32000, '收益为空或0，保底32000'

    # 5. 当前年合同（10月之前，收益不为0）：年化收益×0.3，最低32000
    if year == DATA_YEAR:
        months = 12 - month
        if months <= 0:
            quota = revenue * 0.3
            return max(quota, 32000), f'{DATA_YEAR}年{month}月开设，收益×0.3={quota:.0f}'
        annual_revenue = revenue / months * 12
        quota = annual_revenue * 0.3
        if quota >= 32000:
            return quota, f'{DATA_YEAR}年{month}月开设，{months}个月收益，年化={annual_revenue:.0f}，×0.3={quota:.0f}'
        else:
            return 32000, f'{DATA_YEAR}年{month}月开设，{months}个月收益，年化={annual_revenue:.0f}，×0.3={quota:.0f}<32000，保底32000'

    # 6. 当前年之前合同：收益×0.3，最低32000
    quota = revenue * 0.3
    if quota >= 32000:
        return quota, f'{DATA_YEAR}年前合同，收益×0.3={quota:.0f}'
    else:
        return 32000, f'{DATA_YEAR}年前合同，收益×0.3={quota:.0f}<32000，保底32000'



def main():
    # 文件路径：支持 workflow 时间戳
    ts = os.environ.get('WORKFLOW_TIMESTAMP', '')
    suffix = f'_{ts}' if ts else ''
    input_file = f'中间数据/合同汇总{start_str}-{end_str}{suffix}.xlsx'
    output_file = input_file  # 原地更新

    df = pd.read_excel(input_file)
    print(f'读取: {input_file}, 共 {len(df)} 条记录')

    # 只计算核定总额为空的数据
    mask = df['核定总额'].isna() | (df['核定总额'] == '')

    # 记录原始核定总额有值的数量
    existing_count = len(df) - mask.sum()

    # 计算核定总额和计算规则（仅对空值计算）
    if mask.any():
        df.loc[mask, ['核定总额', '核定总额计算规则']] = df[mask].apply(calc_quota, axis=1, result_type='expand')

    # 用openpyxl原地更新，保留底纹等格式
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # 找到列索引
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_quota = headers.get('核定总额')
    col_rule = headers.get('核定总额计算规则')

    if col_quota and col_rule:
        for row_idx in range(2, ws.max_row + 1):
            quota_val = df.iloc[row_idx - 2]['核定总额']
            rule_val = df.iloc[row_idx - 2]['核定总额计算规则']
            ws.cell(row=row_idx, column=col_quota, value=quota_val)
            ws.cell(row=row_idx, column=col_rule, value=str(rule_val) if pd.notna(rule_val) else '')

    wb.save(output_file)

    print(f'计算完成，结果已保存到: {output_file}')
    print(f'\n统计信息:')
    print(f'总记录数: {len(df)}')
    print(f'已有核定总额（跳过）: {existing_count}条')
    print(f'本次计算核定总额: {mask.sum()}条')
    print(f'核定总额为50000的({NEXT_YEAR}年新开): {len(df[df["核定总额"] == 50000])}条')
    print(f'核定总额为32000的(保底): {len(df[df["核定总额"] == 32000])}条')
    print(f'其他核定总额: {len(df[(df["核定总额"] != 32000) & (df["核定总额"] != 50000)])}条')


if __name__ == '__main__':
    main()

