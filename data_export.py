import sys
import time
import os
import glob
import yaml
import warnings
import openpyxl
import subprocess

sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, timedelta
import calendar

warnings.filterwarnings('ignore', message='Workbook contains no default style')
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pandas as pd


def load_config():
    with open('config.yaml', 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def load_projects():
    """加载项目列表，返回 [(项目名称, URL), ...] 格式"""
    wb = openpyxl.load_workbook('projList.xlsx')
    ws = wb.active
    projects = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        url = ws.cell(row=row, column=2).value
        if name:
            projects.append((name.strip(), url.strip() if url else None))
    return projects


def _kill_chrome_processes():
    """清理残留的 Chrome 和 chromedriver 进程"""
    for proc_name in ['chromedriver.exe', 'chrome.exe']:
        try:
            subprocess.run(['taskkill', '/F', '/IM', proc_name],
                          capture_output=True, timeout=10)
        except Exception:
            pass


def create_chrome_driver(download_dir):
    print("正在初始化 Chrome WebDriver...")

    # 设置 chromedriver 下载镜像（国内加速）
    os.environ['SE_MANAGER_DRIVER_MIRROR_URL'] = 'https://registry.npmmirror.com/-/binary/chrome-for-testing'

    # 清理残留进程
    print("  清理残留 Chrome 进程...")
    _kill_chrome_processes()
    time.sleep(2)

    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--start-maximized')
    options.add_argument('--disable-blink-features=AutomationControlled')

    user_data_dir = os.path.join(os.getcwd(), "chrome_user_data")
    options.add_argument(f'--user-data-dir={user_data_dir}')

    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True
    }
    options.add_experimental_option("prefs", prefs)

    # Selenium Manager 自动下载并缓存 chromedriver（首次运行会下载，之后用缓存）
    print("  启动 Chrome 浏览器...")
    driver = webdriver.Chrome(options=options)
    print("  Chrome WebDriver 初始化完成")
    return driver


def wait_mask_disappear(driver, timeout=10):
    """等待遮罩层消失"""
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, '.mini-mask-background, .mini-mask, .mask'))
        )
    except Exception:
        pass


def wait_and_click(driver, by, value, timeout=10):
    wait_mask_disappear(driver)
    element = WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, value))
    )
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)
    return element


def get_latest_download(download_dir):
    """获取下载目录中最新的Excel文件"""
    excel_files = glob.glob(os.path.join(download_dir, "*.xlsx"))
    excel_files.extend(glob.glob(os.path.join(download_dir, "*.xls")))
    if not excel_files:
        return None
    return max(excel_files, key=os.path.getctime)


def consolidate_to_master(project_name, download_dir, master_file="汇总表.xlsx"):
    """将导出的数据汇总到总表"""
    latest_file = get_latest_download(download_dir)
    if not latest_file:
        print(f"  未找到下载的文件")
        return

    print(f"  读取下载文件: {latest_file}")

    try:
        df = pd.read_excel(latest_file)

        if os.path.exists(master_file):
            with pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                sheet_name = project_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
                sheet_name = project_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"  已添加到总表: {master_file} -> Sheet: {sheet_name}")

        os.remove(latest_file)
        print(f"  已删除原始下载文件")

    except Exception as e:
        print(f"  汇总失败: {e}")


def calc_date_segments(start_date_str, end_date_str, seg_months=3):
    """将日期区间按seg_months个月分段，返回不重叠的(段起始, 段结束)列表"""
    start = date.fromisoformat(start_date_str)
    end = date.fromisoformat(end_date_str)
    segments = []
    seg_start = start
    while seg_start <= end:
        seg_end_candidate = _add_months(seg_start, seg_months) - timedelta(days=1)
        seg_end = min(seg_end_candidate, end)
        segments.append((seg_start.isoformat(), seg_end.isoformat()))
        seg_start = seg_end + timedelta(days=1)
    return segments


def set_date_and_search(driver, seg_start, seg_end):
    """设置日期区间并点击搜索"""
    print(f"    设置日期: {seg_start} ~ {seg_end}")
    wait_mask_disappear(driver, 30)
    try:
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, 'realbegindate$text'))
        )
        driver.execute_script(
            "document.getElementById('realbegindate$value').value = arguments[0];"
            "document.getElementById('realbegindate$text').value = arguments[0];"
            "var obj = mini.get('realbegindate'); if(obj) obj.setValue(arguments[0]);",
            seg_start
        )
        time.sleep(2)

        driver.find_element(By.ID, 'realfinishdate$text')
        driver.execute_script(
            "document.getElementById('realfinishdate$value').value = arguments[0];"
            "document.getElementById('realfinishdate$text').value = arguments[0];"
            "var obj = mini.get('realfinishdate'); if(obj) obj.setValue(arguments[0]);",
            seg_end
        )
        time.sleep(2)
        print(f"    日期设置完成")
    except Exception as e:
        print(f"    日期设置失败: {e}")
        return False

    try:
        print("    点击搜索...")
        wait_mask_disappear(driver, 30)
        search_btn = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.cond-srh-btn-text'))
        )
        try:
            search_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", search_btn)
        time.sleep(10)
        wait_mask_disappear(driver, 60)
        print("    搜索完成")
        return True
    except Exception as e:
        print(f"    搜索失败: {e}")
        return False


def click_export_btn(driver):
    """点击导出Excel按钮，返回是否成功"""
    try:
        print("    点击导出Excel...")
        wait_mask_disappear(driver, 30)
        export_btn = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.ID, 'export'))
        )
        try:
            export_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", export_btn)
        time.sleep(10)
        wait_mask_disappear(driver, 60)
        print("    导出请求已发送")
        return True
    except Exception as e:
        print(f"    导出失败: {e}")
        return False


def wait_for_new_file(download_dir, existing_files, timeout=300, poll_interval=5):
    """轮询等待新文件出现，最长等待timeout秒"""
    print(f"    等待文件下载 (最长{timeout}秒)...")
    elapsed = 0
    while elapsed < timeout:
        time.sleep(poll_interval)
        elapsed += poll_interval
        current_files = set(glob.glob(os.path.join(download_dir, "*.xlsx"))
                           + glob.glob(os.path.join(download_dir, "*.xls")))
        new_files = current_files - existing_files
        if new_files:
            print(f"    文件下载完成 ({elapsed}秒)")
            return new_files.pop()
        # 每30秒打印一次等待状态
        if elapsed % 30 == 0:
            print(f"    已等待 {elapsed} 秒...")
    print(f"    等待超时 ({timeout}秒)")
    return None


def consolidate_segments_to_master(project_name, segment_files, master_file):
    """将多个分段导出文件合并到总表的同一个sheet中"""
    dfs = []
    for f in segment_files:
        try:
            df = pd.read_excel(f)
            if not df.empty:
                dfs.append(df)
        except Exception as e:
            print(f"    读取分段文件失败 {f}: {e}")

    if not dfs:
        print(f"    所有分段均无数据")
        return False

    merged = pd.concat(dfs, ignore_index=True)
    sheet_name = project_name[:31]

    try:
        if os.path.exists(master_file):
            with pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                merged.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
                merged.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"    已合并 {len(dfs)} 个分段到总表 -> Sheet: {sheet_name} (共 {len(merged)} 行)")

        for f in segment_files:
            if os.path.exists(f):
                os.remove(f)
        return True
    except Exception as e:
        print(f"    合并失败: {e}")
        return False


def _add_months(d, months):
    """给日期加上指定月数"""
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def export_by_url(driver, project_name, url, start_date, end_date, download_dir, master_file, split_projects, segment_months):
    """通过直接访问URL进行导出"""
    print(f"  使用URL直接访问: {url}")

    try:
        # 访问URL
        driver.get(url)
        time.sleep(5)
        wait_mask_disappear(driver, 30)

        # 切换到任务列表 iframe
        print("  切换到任务列表 iframe...")
        driver.switch_to.default_content()
        task_iframe_found = False
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
        for idx, iframe in enumerate(iframes):
            try:
                driver.switch_to.frame(iframe)
                test = driver.find_elements(By.ID, 'realbegindate$text')
                if test:
                    print(f"    切换到 iframe[{idx}]，找到日期筛选")
                    task_iframe_found = True
                    break
                driver.switch_to.default_content()
            except Exception:
                driver.switch_to.default_content()

        if not task_iframe_found:
            print("    未找到任务列表 iframe，尝试直接定位...")

        # 判断是否需要分段导出
        need_split = project_name in split_projects

        if need_split:
            # 分段导出模式
            segments = calc_date_segments(start_date, end_date, seg_months=segment_months)
            print(f"  分段导出模式: 每段{segment_months}个月, 共 {len(segments)} 段")
            segment_files = []

            for seg_idx, (seg_start, seg_end) in enumerate(segments, 1):
                print(f"  --- 第 {seg_idx}/{len(segments)} 段: {seg_start} ~ {seg_end} ---")

                # 点击展开更多条件
                try:
                    wait_mask_disappear(driver, 30)
                    expand_btn = WebDriverWait(driver, 60).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, '.cond-srh-btn-toggle'))
                    )
                    try:
                        expand_btn.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", expand_btn)
                    time.sleep(3)
                except Exception as e:
                    print(f"    展开按钮未找到: {e}")

                existing_files = set(glob.glob(os.path.join(download_dir, "*.xlsx"))
                                    + glob.glob(os.path.join(download_dir, "*.xls")))

                ok = set_date_and_search(driver, seg_start, seg_end)
                if not ok:
                    break
                ok = click_export_btn(driver)
                if not ok:
                    break

                new_file = wait_for_new_file(download_dir, existing_files)
                if new_file:
                    segment_files.append(new_file)
                    print(f"    第 {seg_idx} 段导出成功")
                else:
                    print(f"    第 {seg_idx} 段未下载到文件")
                    break

            if segment_files:
                merge_ok = consolidate_segments_to_master(project_name, segment_files, master_file)
                if merge_ok:
                    return {'name': project_name, 'status': 'success',
                            'segments': len(segments), 'detail': 'URL直接访问分段导出成功'}
                else:
                    return {'name': project_name, 'status': 'failed',
                            'segments': len(segments), 'detail': '分段合并失败'}
            else:
                return {'name': project_name, 'status': 'failed',
                        'segments': len(segments), 'detail': '所有分段均无数据'}
        else:
            # 普通导出模式
            print("  点击展开更多条件...")
            try:
                wait_mask_disappear(driver, 30)
                expand_btn = WebDriverWait(driver, 60).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '.cond-srh-btn-toggle'))
                )
                try:
                    expand_btn.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", expand_btn)
                time.sleep(3)
                print("    展开成功")
            except Exception as e:
                print(f"    展开按钮未找到: {e}")

            existing_files = set(glob.glob(os.path.join(download_dir, "*.xlsx"))
                                + glob.glob(os.path.join(download_dir, "*.xls")))
            ok = set_date_and_search(driver, start_date, end_date)
            if ok:
                ok = click_export_btn(driver)
            if ok:
                new_file = wait_for_new_file(download_dir, existing_files)
                if new_file:
                    consolidate_to_master(project_name, download_dir, master_file)
                    return {'name': project_name, 'status': 'success',
                            'segments': 1, 'detail': 'URL直接访问导出成功'}
                else:
                    return {'name': project_name, 'status': 'failed',
                            'segments': 1, 'detail': '导出超时未下载到文件'}
            else:
                return {'name': project_name, 'status': 'failed',
                        'segments': 1, 'detail': '导出失败'}

    except Exception as e:
        print(f"  URL访问失败: {e}")
        return {'name': project_name, 'status': 'failed',
                'segments': 0, 'detail': f'URL访问失败: {e}'}


def main():
    config = load_config()
    projects = load_projects()
    start_date = config['date_range']['start_date']
    end_date = config['date_range']['end_date']

    download_dir = os.path.join(os.getcwd(), "downloads")
    os.makedirs(download_dir, exist_ok=True)

    # 根据日期范围自动生成文件名
    start_str = start_date.replace('-', '')
    end_str = end_date.replace('-', '')
    master_file = f'源数据/export/ejyExport{start_str}-{end_str}.xlsx'
    os.makedirs('源数据/export', exist_ok=True)
    if os.path.exists(master_file):
        os.remove(master_file)

    split_config = config.get('split_export', {})
    split_projects = split_config.get('projects', []) if isinstance(split_config, dict) else split_config
    segment_months = split_config.get('segment_months', 3) if isinstance(split_config, dict) else 3
    export_report = []
    project_urls = {}

    # 统计有URL和无URL的项目
    projects_with_url = [(name, url) for name, url in projects if url]
    projects_without_url = [(name, None) for name, url in projects if not url]

    print(f"加载了 {len(projects)} 个项目")
    print(f"  有URL: {len(projects_with_url)} 个")
    print(f"  无URL: {len(projects_without_url)} 个")
    print(f"日期范围: {start_date} ~ {end_date}")
    print(f"下载目录: {download_dir}")
    print(f"汇总文件: {master_file}")

    driver = create_chrome_driver(download_dir)
    wait = WebDriverWait(driver, 15)

    try:
        url = "https://oa.epoint.com.cn/epoint-projectmanage-web/frame/fui/pages/themes/grace/grace?pageId=grace"
        print(f"正在访问: {url}")
        driver.get(url)
        print("页面加载完成")

        user_data_dir = os.path.join(os.getcwd(), "chrome_user_data")
        if os.path.exists(user_data_dir):
            print("检测到已保存的登录态，等待页面加载...")
            time.sleep(5)
        else:
            print("首次运行，请扫码登录...")
            print("登录态将自动保存，下次无需重新登录")

        print("等待登录完成...")
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'li[data-id="00050007"]'))
        )
        print("登录成功！")

        # 第一步：先处理无URL的项目（使用搜索方式）
        if projects_without_url:
            print(f"\n{'='*60}")
            print(f"第一步：处理无URL的项目 ({len(projects_without_url)} 个)")
            print(f"{'='*60}")

            print("点击项目集菜单...")
            wait_and_click(driver, By.CSS_SELECTOR, 'li[data-id="00050007"]')
            print("项目集菜单点击完成，等待页面加载...")
            time.sleep(5)

            wait_mask_disappear(driver, 15)

            print("查找并切换到 iframe...")
            driver.switch_to.default_content()

            switched = False
            for attempt in range(3):
                print(f"  第 {attempt+1} 次尝试查找 iframe...")
                time.sleep(3)
                iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                print(f"  找到 {len(iframes)} 个 iframe")
                for idx, iframe in enumerate(iframes):
                    try:
                        driver.switch_to.frame(iframe)
                        test = driver.find_elements(By.CSS_SELECTOR, 'input#xiangmumc\\$text')
                        if test:
                            print(f"  已切换到 iframe[{idx}]，找到搜索框")
                            switched = True
                            break
                        driver.switch_to.default_content()
                    except Exception:
                        driver.switch_to.default_content()
                if switched:
                    break

            if not switched:
                print("  未找到包含搜索框的 iframe，尝试直接定位...")

            for i, (project_name, _) in enumerate(projects_without_url, 1):
                print(f"\n处理第 {i}/{len(projects_without_url)} 个项目: {project_name}")

                if i == 1:
                    print("  首次加载，额外等待页面完全就绪...")
                    time.sleep(10)
                    wait_mask_disappear(driver, 30)

                print("  等待搜索框加载...")
                try:
                    search_box = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'input#xiangmumc\\$text'))
                    )
                except Exception:
                    print("  搜索框未找到，重新切换 iframe...")
                    driver.switch_to.default_content()
                    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                    for idx, iframe in enumerate(iframes):
                        try:
                            driver.switch_to.frame(iframe)
                            test = driver.find_elements(By.CSS_SELECTOR, 'input#xiangmumc\\$text')
                            if test:
                                break
                            driver.switch_to.default_content()
                        except Exception:
                            driver.switch_to.default_content()
                    search_box = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'input#xiangmumc\\$text'))
                    )

                print("  搜索框已找到，输入项目名称...")
                search_box.clear()
                search_box.send_keys(project_name)
                time.sleep(1)

                print("  点击搜索...")
                wait_mask_disappear(driver)
                search_btn = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '.cond-srh-btn-text'))
                )
                try:
                    search_btn.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", search_btn)
                time.sleep(5)
                wait_mask_disappear(driver, 15)

                print("  点击第一条搜索结果进入详情...")
                try:
                    wait_mask_disappear(driver, 15)
                    time.sleep(2)
                    first_project = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, '.mini-grid-row .mini-tree-nodetext i[onclick*="openXMGL"]'))
                    )
                    main_window = driver.current_window_handle
                    try:
                        first_project.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", first_project)
                    time.sleep(5)

                    handles = driver.window_handles
                    if len(handles) > 1:
                        for handle in handles:
                            if handle != main_window:
                                driver.switch_to.window(handle)
                                break

                    print("  已进入详情页")
                    print(f"  详情页标题: {driver.title}")

                    print("  点击任务菜单...")
                    try:
                        wait_mask_disappear(driver)
                        task_link = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.ID, '任务'))
                        )
                        task_href = task_link.get_attribute('href')
                        if task_href:
                            project_urls[project_name] = task_href
                            print(f"  获取链接: {task_href}")
                        try:
                            task_link.click()
                        except Exception:
                            driver.execute_script("arguments[0].click();", task_link)
                        time.sleep(5)
                        print("  已进入任务列表")

                        print("  切换到任务列表 iframe...")
                        driver.switch_to.default_content()
                        task_iframe_found = False
                        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                        for idx, iframe in enumerate(iframes):
                            try:
                                driver.switch_to.frame(iframe)
                                test = driver.find_elements(By.ID, 'realbegindate$text')
                                if test:
                                    print(f"    切换到 iframe[{idx}]，找到日期筛选")
                                    task_iframe_found = True
                                    break
                                driver.switch_to.default_content()
                            except Exception:
                                driver.switch_to.default_content()

                        if not task_iframe_found:
                            print("    未找到任务列表 iframe，尝试直接定位...")

                        # 判断是否需要分段导出（只要在配置列表中就分段）
                        need_split = project_name in split_projects

                        if need_split:
                            # 分段导出模式：先展开条件，再按段循环
                            segments = calc_date_segments(start_date, end_date, seg_months=segment_months)
                            print(f"  分段导出模式: 每段{segment_months}个月, 共 {len(segments)} 段")
                            segment_files = []

                            for seg_idx, (seg_start, seg_end) in enumerate(segments, 1):
                                print(f"  --- 第 {seg_idx}/{len(segments)} 段: {seg_start} ~ {seg_end} ---")

                                # 点击展开更多条件（每段都需要重新展开）
                                try:
                                    wait_mask_disappear(driver, 30)
                                    expand_btn = WebDriverWait(driver, 60).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, '.cond-srh-btn-toggle'))
                                    )
                                    try:
                                        expand_btn.click()
                                    except Exception:
                                        driver.execute_script("arguments[0].click();", expand_btn)
                                    time.sleep(3)
                                except Exception as e:
                                    print(f"    展开按钮未找到: {e}")

                                existing_files = set(glob.glob(os.path.join(download_dir, "*.xlsx"))
                                                    + glob.glob(os.path.join(download_dir, "*.xls")))

                                ok = set_date_and_search(driver, seg_start, seg_end)
                                if not ok:
                                    break
                                ok = click_export_btn(driver)
                                if not ok:
                                    break

                                new_file = wait_for_new_file(download_dir, existing_files)
                                if new_file:
                                    segment_files.append(new_file)
                                    print(f"    第 {seg_idx} 段导出成功")
                                else:
                                    print(f"    第 {seg_idx} 段未下载到文件")
                                    break

                            if segment_files:
                                merge_ok = consolidate_segments_to_master(project_name, segment_files, master_file)
                                if merge_ok:
                                    export_report.append({'name': project_name, 'status': 'success',
                                                         'segments': len(segments), 'detail': '分段导出成功'})
                                else:
                                    export_report.append({'name': project_name, 'status': 'failed',
                                                         'segments': len(segments), 'detail': '分段合并失败'})
                            else:
                                export_report.append({'name': project_name, 'status': 'failed',
                                                     'segments': len(segments), 'detail': '所有分段均无数据'})
                        else:
                            # 普通导出模式
                            print("  点击展开更多条件...")
                            try:
                                wait_mask_disappear(driver, 30)
                                expand_btn = WebDriverWait(driver, 60).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '.cond-srh-btn-toggle'))
                                )
                                try:
                                    expand_btn.click()
                                except Exception:
                                    driver.execute_script("arguments[0].click();", expand_btn)
                                time.sleep(3)
                                print("    展开成功")
                            except Exception as e:
                                print(f"    展开按钮未找到: {e}")

                            existing_files = set(glob.glob(os.path.join(download_dir, "*.xlsx"))
                                                + glob.glob(os.path.join(download_dir, "*.xls")))
                            ok = set_date_and_search(driver, start_date, end_date)
                            if ok:
                                ok = click_export_btn(driver)
                            if ok:
                                new_file = wait_for_new_file(download_dir, existing_files)
                                if new_file:
                                    consolidate_to_master(project_name, download_dir, master_file)
                                    export_report.append({'name': project_name, 'status': 'success',
                                                         'segments': 1, 'detail': '普通导出成功'})
                                else:
                                    export_report.append({'name': project_name, 'status': 'failed',
                                                         'segments': 1, 'detail': '导出超时未下载到文件'})
                            else:
                                export_report.append({'name': project_name, 'status': 'failed',
                                                     'segments': 1, 'detail': '导出失败'})

                    except Exception as e:
                        print(f"  任务操作失败: {e}")
                        export_report.append({'name': project_name, 'status': 'failed',
                                             'segments': 0, 'detail': f'任务操作失败: {e}'})

                    print("  关闭详情页标签...")
                    try:
                        handles = driver.window_handles
                        if len(handles) > 1:
                            driver.close()
                        driver.switch_to.window(main_window)
                    except Exception:
                        try:
                            driver.switch_to.window(main_window)
                        except Exception:
                            pass
                    time.sleep(2)

                    print("  重新切换到主列表 iframe...")
                    driver.switch_to.default_content()
                    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                    for idx, iframe in enumerate(iframes):
                        try:
                            driver.switch_to.frame(iframe)
                            test = driver.find_elements(By.CSS_SELECTOR, 'input#xiangmumc\\$text')
                            if test:
                                print(f"    已切换到 iframe[{idx}]")
                                break
                            driver.switch_to.default_content()
                        except Exception:
                            driver.switch_to.default_content()

                except Exception as e:
                    print(f"  点击第一条结果失败: {e}")
                    export_report.append({'name': project_name, 'status': 'failed',
                                         'segments': 0, 'detail': f'进入详情页失败: {e}'})

        # 第二步：处理有URL的项目（直接访问URL）
        if projects_with_url:
            print(f"\n{'='*60}")
            print(f"第二步：处理有URL的项目 ({len(projects_with_url)} 个)")
            print(f"{'='*60}")

            for i, (project_name, project_url) in enumerate(projects_with_url, 1):
                print(f"\n处理第 {i}/{len(projects_with_url)} 个项目: {project_name}")

                result = export_by_url(driver, project_name, project_url, start_date, end_date,
                                      download_dir, master_file, split_projects, segment_months)
                export_report.append(result)

                # 回到主页准备下一个项目
                if i < len(projects_with_url):
                    print("  返回主页...")
                    driver.get("https://oa.epoint.com.cn/epoint-projectmanage-web/frame/fui/pages/themes/grace/grace?pageId=grace")
                    time.sleep(3)

    except Exception as e:
        print(f"发生错误: {e}")
    finally:
        print("\n所有项目处理完成，10秒后关闭浏览器...")
        time.sleep(10)
        driver.quit()

        # 输出导出报告
        print("\n" + "=" * 60)
        print("导出报告")
        print("=" * 60)
        success_count = 0
        fail_count = 0
        for r in export_report:
            status_icon = "[OK]" if r['status'] == 'success' else "[FAIL]"
            seg_info = f"({r['segments']}段)" if r['segments'] > 1 else ""
            print(f"  {status_icon} {r['name']} {seg_info} - {r['detail']}")
            if r['status'] == 'success':
                success_count += 1
            else:
                fail_count += 1

        if not export_report:
            print("  (无项目被处理)")
        else:
            print(f"\n合计: 成功 {success_count}, 失败 {fail_count}, 共 {len(export_report)} 个专区")

        if fail_count > 0:
            print("\n以下专区导出失败，请检查:")
            for r in export_report:
                if r['status'] != 'success':
                    print(f"  - {r['name']}: {r['detail']}")

        if os.path.exists(master_file):
            wb = openpyxl.load_workbook(master_file)
            print(f"\n汇总文件: {master_file}")
            print(f"共 {len(wb.sheetnames)} 个Sheet: {', '.join(wb.sheetnames)}")

        # 回写链接到 projList.xlsx
        if project_urls:
            wb_proj = openpyxl.load_workbook('projList.xlsx')
            ws_proj = wb_proj.active
            ws_proj.cell(row=1, column=2, value='链接')
            updated_count = 0
            for row in range(2, ws_proj.max_row + 1):
                name = ws_proj.cell(row=row, column=1).value
                if name and name.strip() in project_urls:
                    ws_proj.cell(row=row, column=2, value=project_urls[name.strip()])
                    updated_count += 1
            wb_proj.save('projList.xlsx')
            print(f"\n已回写 {updated_count} 个链接到 projList.xlsx")


if __name__ == "__main__":
    main()
