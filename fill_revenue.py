"""
收益回填脚本：读取下载的收益文件，根据辖区code匹配流程一输出的专区码，回填26年收益列
"""
import pandas as pd
import openpyxl
import os
import yaml
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')


def load_config():
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.yaml'), 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def main():
    config = load_config()
    start_date = config['date_range']['start_date']
    end_date = config['date_range']['end_date']
    start_str = start_date.replace('-', '')
    end_str = end_date.replace('-', '')

    # 数据年份：优先使用yaml配置，否则根据当前年份自动生成
    _config_year = config.get('dataYear', '')
    if _config_year:
        data_year_short = _config_year
    else:
        data_year_short = str(datetime.now().year)[2:]
    year_revenue_col = f'{data_year_short}年收益'

    base_dir = os.path.dirname(os.path.abspath(__file__))
    revenue_file = os.path.join(base_dir, f'源数据/export/shouyi{start_str}-{end_str}.xlsx')

    # 支持 workflow 时间戳
    ts = os.environ.get('WORKFLOW_TIMESTAMP', '')
    suffix = f'_{ts}' if ts else ''
    target_file = os.path.join(base_dir, f'中间数据/合同汇总{start_str}-{end_str}{suffix}.xlsx')

    print(f"收益文件: {revenue_file}")
    print(f"流程一输出: {target_file}")
    print(f"收益列名: {year_revenue_col}")

    if not os.path.exists(revenue_file):
        print(f"收益文件不存在: {revenue_file}")
        return False

    if not os.path.exists(target_file):
        print(f"流程一输出文件不存在: {target_file}")
        return False

    # 读取收益文件，构建 辖区code -> 实得收益 映射
    df_revenue = pd.read_excel(revenue_file)
    print(f"读取收益文件: 共 {len(df_revenue)} 条记录")

    if '辖区code' not in df_revenue.columns or '实得收益' not in df_revenue.columns:
        print(f"收益文件缺少必要列。当前列: {list(df_revenue.columns)}")
        return False

    revenue_map = {}
    for _, row in df_revenue.iterrows():
        code = str(row['辖区code']).strip() if pd.notna(row['辖区code']) else ''
        revenue = row['实得收益'] if pd.notna(row['实得收益']) else 0
        if code and code != 'nan':
            revenue_map[code] = revenue_map.get(code, 0) + revenue
    print(f"收益映射数量: {len(revenue_map)}")

    # 用openpyxl原地更新，保留底纹等格式
    wb = openpyxl.load_workbook(target_file)
    ws = wb.active

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # 如果收益列不存在，新增
    if year_revenue_col not in headers:
        next_col = ws.max_column + 1
        ws.cell(row=1, column=next_col, value=year_revenue_col)
        headers[year_revenue_col] = next_col
        print(f"新增列: {year_revenue_col}")

    col_zone = headers.get('专区码')
    col_revenue = headers.get(year_revenue_col)

    if not col_zone:
        print("流程一输出文件缺少'专区码'列")
        wb.close()
        return False

    # 匹配并回填
    filled_count = 0
    for row_idx in range(2, ws.max_row + 1):
        zone_code = str(ws.cell(row=row_idx, column=col_zone).value).strip() if ws.cell(row=row_idx, column=col_zone).value else ''
        if zone_code in revenue_map:
            ws.cell(row=row_idx, column=col_revenue, value=revenue_map[zone_code])
            filled_count += 1

    wb.save(target_file)
    print(f"\n回填完成: {target_file}")
    print(f"匹配成功: {filled_count} / {ws.max_row - 1} 行")
    return True


if __name__ == '__main__':
    main()
