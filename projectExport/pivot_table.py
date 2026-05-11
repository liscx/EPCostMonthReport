import os
import pandas as pd
from openpyxl import load_workbook

FILE = '源数据/export/ejyExport.xlsx'
OUTPUT = None  # None则覆盖原文件

# 数据透视表
def create_pivot_table(df, project_name):
    """为单个项目创建数据透视表"""
    required_cols = ['合同编号', '任务类型', '任务预算使用']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"  缺少列: {missing_cols}")
        print(f"  现有列: {list(df.columns)}")
        return None

    df['任务预算使用'] = pd.to_numeric(df['任务预算使用'], errors='coerce').fillna(0)
    df['合同编号'] = df['合同编号'].fillna('(空白)')
    df['任务类型'] = df['任务类型'].fillna('(空白)')

    pivot = pd.pivot_table(
        df,
        values='任务预算使用',
        index='任务类型',
        columns='合同编号',
        aggfunc='sum',
        fill_value=0,
        margins=True,
        margins_name='总计'
    )

    return pivot


def main():
    if not os.path.exists(FILE):
        print(f"文件 {FILE} 不存在")
        return

    wb = load_workbook(FILE)
    sheet_names = wb.sheetnames
    print(f"共 {len(sheet_names)} 个项目待分析")

    for i, sheet_name in enumerate(sheet_names, 1):
        if sheet_name.endswith('-数据透视'):
            continue

        print(f"\n处理第 {i}/{len(sheet_names)} 个项目: {sheet_name}")

        try:
            df = pd.read_excel(FILE, sheet_name=sheet_name)

            if df.empty:
                print(f"  数据为空，跳过")
                continue

            pivot = create_pivot_table(df, sheet_name)

            if pivot is not None:
                pivot_sheet_name = f"{sheet_name[:28]}-数据透视"

                if pivot_sheet_name in wb.sheetnames:
                    del wb[pivot_sheet_name]

                ws = wb.create_sheet(pivot_sheet_name)

                ws.cell(row=1, column=1, value='求和项:任务预算使用')
                ws.cell(row=1, column=2, value='合同编号')

                ws.cell(row=2, column=1, value='任务类型')
                for c_idx, col_name in enumerate(pivot.columns, 2):
                    ws.cell(row=2, column=c_idx, value=col_name)

                for r_idx, (idx_name, row) in enumerate(pivot.iterrows(), 3):
                    ws.cell(row=r_idx, column=1, value=idx_name)
                    for c_idx, value in enumerate(row, 2):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                print(f"  透视表已生成: {pivot_sheet_name}")
                print(f"  行数: {len(pivot)}, 列数: {len(pivot.columns)}")

        except Exception as e:
            print(f"  处理失败: {e}")

    output = OUTPUT if OUTPUT is not None else FILE
    wb.save(output)
    print(f"\n分析完成！已更新文件: {output}")


if __name__ == "__main__":
    main()
